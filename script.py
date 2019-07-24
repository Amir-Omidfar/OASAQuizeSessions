#To write to excel files
from openpyxl import Workbook, load_workbook

quiz_file=load_workbook("2019 TR Quiz Check.xlsx")
quizSheetOutput=quiz_file["Sheet1"]




# Reading an excel file
import xlrd 


#Give  the location of the file
quizLoc = ("2019 TR Quiz Check.xlsx")
csLoc = ("CS_2019_orientation2.xlsx")

#To open Workbook

#This is our final place to gather all the information regardingn students' quizes 
quizWb = xlrd.open_workbook(quizLoc)
quizSheet = quizWb.sheet_by_index(0)

#These are the spread sheets for each major that we use to update our final code
#cssheet for CS student
csWb = xlrd.open_workbook(csLoc)
csSheet = csWb.sheet_by_index(0)




#A dictionary is used here to go over all the entries in our list and see which student is missing the quiz
dict={}
address={}


for i in range(quizSheet.nrows):
	if (quizSheet.cell_value(i,4) == ""):
		dict[quizSheet.cell_value(i,0)]=quizSheet.cell_value(i,1)
		address[quizSheet.cell_value(i,0)]=i
print(dict)
print(len(dict))
#Then below lists will have the UID's based on their majors so we can look up the correct excel file to update our 
#Form
CSList=[]
#EEList=[]

for i in dict:
	if (dict[i] == "COM SCI") or (dict[i] == "C S&ENGR") :
		CSList.append(i)

print(CSList)

#quizSheet['J1']=""

for i in range(csSheet.nrows):
	if(csSheet.cell_value(i,5) in CSList):
		quizSession=csSheet.cell_value(i,1)
		quizSheetOutput.cell(address[csSheet.cell_value(i,5)]+1,5,quizSession)


quiz_file.save("2019 TR Quiz Check modified.xlsx")





