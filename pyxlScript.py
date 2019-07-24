from openpyxl import Workbook, load_workbook

quiz_file=load_workbook("2019 TR Quiz Check.xlsx")
#Majors files 
cs_file=load_workbook("CS_2019_orientation.xlsx")

quizSheet=quiz_file["Sheet1"]
csSheet=cs_file["report"]

#print(quizSheet['A3'].value)
#quizSheet['J1']=""

#A dictionary is used here to go over all the entries in our list and see which student is missing the quiz
dict={}
address={}


for col in quizSheet.iter_cols(min_row=1,max_row=247,min_col=5,max_col=5):
	for cell in col:
		if (cell.value == ""):
			dict{cell(1,)}

'''
for i in range(quizSheet.nrows):
	if (quizSheet.cell_value(i,4) == ""):
		dict[quizSheet.cell_value(i,0)]=quizSheet.cell_value(i,1)
		address[quizSheet.cell_value(i,0)]=i

'''

#Save the files after changes 
quiz_file.save("2019 TR Quiz Check modified.xlsx")