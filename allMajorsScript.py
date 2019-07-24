#Here we implement all the majors in the same file
#To write to excel files
from openpyxl import Workbook, load_workbook
quiz_file=load_workbook("2019 TR Quiz Check.xlsx")
quizSheetOutput=quiz_file["Sheet1"]
# Reading an excel file
import xlrd 




#Give  the location of the Quiz sheet
quizLoc = ("2019 TR Quiz Check.xlsx")

#Give the location of the remaining files
aeroLoc = ("AttendeesList/OrAttendByMajor AeroSpace.xlsx")
bioLoc = ("AttendeesList/OrAttendByMajor Bio.xlsx")
ceLoc = ("AttendeesList/OrAttendByMajor CE.xlsx")
civilLoc = ("AttendeesList/OrAttendByMajor Civil.xlsx")
csLoc = ("AttendeesList/OrAttendByMajor CS.xlsx")
cseLoc = ("AttendeesList/OrAttendByMajor CSE.xlsx")
eeLoc = ("AttendeesList/OrAttendByMajor EE.xlsx")
matLoc = ("AttendeesList/OrAttendByMajor MAT.xlsx")
meLoc = ("AttendeesList/OrAttendByMajor ME.xlsx")

#To open Workbook

#This is our final place to gather all the information regardingn students' quizes 
quizWb = xlrd.open_workbook(quizLoc)
quizSheet = quizWb.sheet_by_index(0)

#These are the spread sheets for each major that we use to update our final code
#aeroSheet for AeroSpace students
aeroWb = xlrd.open_workbook(aeroLoc)
aeroSheet = aeroWb.sheet_by_index(0)
#bioSheet for BioEngineering students
bioWb = xlrd.open_workbook(bioLoc)
bioSheet = bioWb.sheet_by_index(0)
#ceSheet for Computer Engineering students
ceWb = xlrd.open_workbook(ceLoc)
ceSheet = ceWb.sheet_by_index(0)
#civilSheet for Civil Engineering students
civilWb = xlrd.open_workbook(civilLoc)
civilSheet = civilWb.sheet_by_index(0)
#cssheet for Com Science student
csWb = xlrd.open_workbook(csLoc)
csSheet = csWb.sheet_by_index(0)
#cseheet for Computer Science and Engineering students
cseWb = xlrd.open_workbook(cseLoc)
cseSheet = cseWb.sheet_by_index(0)
#eesheet for Electrical Engineering student
eeWb = xlrd.open_workbook(eeLoc)
eeSheet = eeWb.sheet_by_index(0)
#matsheet for Materials Science students
matWb = xlrd.open_workbook(matLoc)
matSheet = matWb.sheet_by_index(0)
#mesheet for Mechanical Engineering students
meWb = xlrd.open_workbook(meLoc)
meSheet = meWb.sheet_by_index(0)

#A dictionary is used here to go over all the entries in our list and see which student is missing the quiz
dict={}
address={}


for i in range(quizSheet.nrows):
	if (quizSheet.cell_value(i,4) == ""):
		dict[quizSheet.cell_value(i,0)]=quizSheet.cell_value(i,1)
		address[quizSheet.cell_value(i,0)]=i
print(dict)
print(len(dict))

CSList=[]
CSEList=[]
AEROList=[]
EEList=[]
BIOList=[]
CEList=[]
CHMList=[]
CIVList=[]
MATList=[]
MEList=[]


#Counting the outputs of each file for assuring the results are correct

for i in dict:
	if (dict[i] == "COM SCI"):
		CSList.append(i)
	elif (dict[i] == "C S&ENGR"):
		CSEList.append(i)
	elif (dict[i] == "AEROSPCE"):
		AEROList.append(i)
	elif (dict[i] == "BIOENGR"):
		BIOList.append(i)
	elif (dict[i] == "COM ENGR"):
		CEList.append(i)
	elif (dict[i] == "CHM ENGR"):
		CHMList.append(i)
	elif (dict[i] == "CIV ENGR"):
		CIVList.append(i)
	elif (dict[i] == "ELE ENGR"):
		EEList.append(i)
	elif (dict[i] == "MAT ENGR"):
		MATList.append(i)
	elif(dict[i] == "MECHANIC"):
		MEList.append(i)


print("AEROSPCE",AEROList,len(AEROList))
print("BIOENGR",BIOList,len(BIOList))
print("CE",CEList,len(CEList))
print("CHM",CHMList,len(CHMList))
print("Civil",CIVList,len(CIVList))
print("CS",CSList,len(CSList))
print("CSE",CSEList,len(CSEList))
print("EE",EEList,len(EEList))
print("MAT",MATList,len(MATList))
print("ME",MEList,len(MEList))


#Checking for the quiz sessions and fill up the Quiz form

for i in range(csSheet.nrows):
	if(csSheet.cell_value(i,5) in CSList):
		quizSession=csSheet.cell_value(i,1)
		quizSheetOutput.cell(address[csSheet.cell_value(i,5)]+1,5,quizSession)


for i in range(eeSheet.nrows):
	if(eeSheet.cell_value(i,5) in EEList):
		quizSession=eeSheet.cell_value(i,1)
		quizSheetOutput.cell(address[eeSheet.cell_value(i,5)]+1,5,quizSession)

for i in range(aeroSheet.nrows):
	if(aeroSheet.cell_value(i,5) in AEROList):
		quizSession=aeroSheet.cell_value(i,1)
		quizSheetOutput.cell(address[aeroSheet.cell_value(i,5)]+1,5,quizSession)

for i in range(civilSheet.nrows):
	if(civilSheet.cell_value(i,5) in CIVList):
		quizSession=civilSheet.cell_value(i,1)
		quizSheetOutput.cell(address[civilSheet.cell_value(i,5)]+1,5,quizSession)


quiz_file.save("2019 TR Quiz Check updatedNoChem.xlsx")










